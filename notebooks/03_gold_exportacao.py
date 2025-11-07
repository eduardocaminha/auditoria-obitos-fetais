# Databricks notebook source
# MAGIC %md
# MAGIC # Gold: Consolidação e Exportação
# MAGIC 
# MAGIC Gera planilha Excel com resultados consolidados a partir da Silver.
# MAGIC Notebook desenhado para execução manual.

# COMMAND ----------

from pyspark.sql import functions as F
from pyspark.sql.utils import AnalysisException
import pandas as pd
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# COMMAND ----------

# MAGIC %md
# MAGIC ## 1. Configuração

# COMMAND ----------

SILVER_TABLE = "innovation_dev.silver.auditoria_obitos_fetais_processado"

# Ajuste o path conforme seu ambiente Databricks
OUTPUT_PATH = "/Workspace/Innovation/t_eduardo.caminha/auditoria-obitos-fetais/outputs"
# Alternativa: usar DBFS
# OUTPUT_PATH = "/dbfs/FileStore/auditoria-obitos-fetais/outputs"

FILTRAR_POR_PERIODO = True
PERIODO_INICIO = '2025-10-01'
PERIODO_FIM = '2025-11-01'

DATA_PROCESSAMENTO = datetime.now().strftime('%Y%m%d_%H%M%S')
EXCEL_FILENAME = f"obitos_fetais_{DATA_PROCESSAMENTO}.xlsx"
EXCEL_PATH = f"{OUTPUT_PATH}/{EXCEL_FILENAME}"

print("=" * 80)
print("CONFIGURAÇÃO GOLD")
print("=" * 80)
print(f"Tabela Silver: {SILVER_TABLE}")
print(f"Diretório de saída: {OUTPUT_PATH}")
print(f"Arquivo: {EXCEL_FILENAME}")
if FILTRAR_POR_PERIODO:
    print(f"Período filtrado: {PERIODO_INICIO} a {PERIODO_FIM}")
else:
    print("Período filtrado: NÃO (todos os dados da Silver)")
print("=" * 80)

# COMMAND ----------

# MAGIC %md
# MAGIC ## 2. Ler Silver

# COMMAND ----------

try:
    silver_df = spark.table(SILVER_TABLE)
    print(f"✅ Tabela Silver encontrada")
except AnalysisException:
    raise RuntimeError(f"⚠️ Tabela Silver {SILVER_TABLE} não encontrada. Execute o processamento da Silver antes do Gold.")

# Aplicar filtro de período se necessário
if FILTRAR_POR_PERIODO:
    silver_df = silver_df.filter(
        (F.col('dt_procedimento_realizado') >= F.to_timestamp(F.lit(PERIODO_INICIO))) &
        (F.col('dt_procedimento_realizado') < F.to_timestamp(F.lit(PERIODO_FIM)))
    )
    print(f"📅 Filtro de período aplicado: {PERIODO_INICIO} a {PERIODO_FIM}")

total_registros = silver_df.count()
print(f"📊 Registros carregados da Silver: {total_registros:,}")

if total_registros == 0:
    raise RuntimeError("⚠️ Nenhum registro encontrado para exportação. Verifique o filtro de período.")

silver_pd = silver_df.toPandas()

# COMMAND ----------

# MAGIC %md
# MAGIC ## 3. Preparar DataFrames

# COMMAND ----------

# Nota: Silver já contém APENAS óbitos fetais detectados (obito_fetal_clinico == 1)
silver_pd['dt_procedimento_realizado'] = pd.to_datetime(silver_pd['dt_procedimento_realizado'])

# DataFrame com pacientes únicos (primeiro óbito detectado)
obitos_unicos_pd = silver_pd.drop_duplicates(subset=['cd_paciente'], keep='first').copy()

print(f"📋 Total de laudos com óbito fetal: {len(silver_pd):,}")
print(f"👥 Pacientes únicos com óbito fetal: {len(obitos_unicos_pd):,}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 4. Estatísticas

# COMMAND ----------

total_obitos = len(silver_pd)
pacientes_unicos = len(obitos_unicos_pd)

# Estatísticas por fonte (HSP vs PSC)
stats_fonte = (
    silver_pd
    .groupby('fonte')
    .size()
    .reset_index(name='total_laudos')
)
stats_fonte['pacientes_unicos'] = silver_pd.groupby('fonte')['cd_paciente'].nunique().values

# Estatísticas por termo detectado
stats_termo = (
    silver_pd['termo_detectado']
    .value_counts(dropna=False)
    .reset_index()
    .rename(columns={'termo_detectado': 'quantidade'})
)
stats_termo['percentual'] = (stats_termo['quantidade'] / total_obitos * 100).round(2)

# Estatísticas por período (mês)
silver_pd['mes_ano'] = silver_pd['dt_procedimento_realizado'].dt.to_period('M').astype(str)
stats_periodo = (
    silver_pd
    .groupby('mes_ano')
    .agg(
        total_laudos=('cd_atendimento', 'count'),
        pacientes_unicos=('cd_paciente', 'nunique')
    )
    .reset_index()
    .sort_values('mes_ano')
)

# Overview geral
stats_overview = pd.DataFrame({
    'metrica': [
        'Total de laudos com óbito fetal',
        'Pacientes únicos com óbito fetal',
        'Média de laudos por paciente',
        'Fonte HSP',
        'Fonte PSC',
        'Período analisado',
        'Data de processamento'
    ],
    'valor': [
        total_obitos,
        pacientes_unicos,
        f"{(total_obitos / pacientes_unicos):.2f}" if pacientes_unicos > 0 else "0",
        stats_fonte[stats_fonte['fonte'] == 'HSP']['total_laudos'].values[0] if 'HSP' in stats_fonte['fonte'].values else 0,
        stats_fonte[stats_fonte['fonte'] == 'PSC']['total_laudos'].values[0] if 'PSC' in stats_fonte['fonte'].values else 0,
        f"{PERIODO_INICIO} a {PERIODO_FIM}" if FILTRAR_POR_PERIODO else "Todos os dados",
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ]
})

print("✅ Estatísticas calculadas")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 5. Preparar Dados para Exportação

# COMMAND ----------

colunas_export = [
    'fonte',
    'cd_atendimento',
    'cd_paciente',
    'nm_paciente',
    'cd_procedimento',
    'nm_procedimento',
    'dt_procedimento_realizado',
    'termo_detectado',
    'texto_original'
]

# Todos os óbitos detectados
obitos_export = silver_pd[colunas_export].copy()
obitos_export = obitos_export.sort_values('dt_procedimento_realizado', ascending=False)

# Pacientes únicos (primeiro óbito)
obitos_unicos_export = obitos_unicos_pd[colunas_export].copy()
obitos_unicos_export = obitos_unicos_export.sort_values('dt_procedimento_realizado', ascending=False)

print(f"✅ Dados preparados para exportação: {len(obitos_export):,} laudos")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 6. Gerar Excel

# COMMAND ----------

# Criar workbook
wb = Workbook()
wb.remove(wb.active)

def escrever_aba(nome, dataframe):
    """Escreve um DataFrame em uma aba do Excel com formatação"""
    ws = wb.create_sheet(nome)
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        ws.append(row)
    
    # Formatar cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # Ajustar largura das colunas
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(adjusted_width, 60)

# Criar abas
escrever_aba("1_Resumo", stats_overview)
escrever_aba("2_Todos_Obitos", obitos_export)
escrever_aba("3_Pacientes_Unicos", obitos_unicos_export)
escrever_aba("4_Por_Fonte", stats_fonte)
escrever_aba("5_Por_Termo", stats_termo)
escrever_aba("6_Por_Periodo", stats_periodo)

# Salvar arquivo
import os
os.makedirs(OUTPUT_PATH, exist_ok=True)
wb.save(EXCEL_PATH)

print("=" * 80)
print("✅ ARQUIVO EXCEL GERADO COM SUCESSO!")
print("=" * 80)
print(f"📁 Local: {EXCEL_PATH}")
print(f"📊 Abas criadas:")
print(f"   1. Resumo Geral")
print(f"   2. Todos os Óbitos ({len(obitos_export):,} laudos)")
print(f"   3. Pacientes Únicos ({len(obitos_unicos_export):,} pacientes)")
print(f"   4. Estatísticas por Fonte")
print(f"   5. Estatísticas por Termo")
print(f"   6. Estatísticas por Período")
print("=" * 80)

# COMMAND ----------

# MAGIC %md
# MAGIC ## 7. Resumo Final

# COMMAND ----------

print("=" * 80)
print("RESUMO DA EXPORTAÇÃO GOLD")
print("=" * 80)
print(f"📊 Total de óbitos fetais: {total_obitos:,}")
print(f"👥 Pacientes únicos: {pacientes_unicos:,}")
print(f"📅 Período: {PERIODO_INICIO} a {PERIODO_FIM}" if FILTRAR_POR_PERIODO else "📅 Período: Todos os dados")
print(f"📁 Arquivo: {EXCEL_FILENAME}")
print(f"✅ Status: Exportação concluída com sucesso")
print("=" * 80)


