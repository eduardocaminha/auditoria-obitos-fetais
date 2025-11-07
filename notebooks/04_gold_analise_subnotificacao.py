# Databricks notebook source
# MAGIC %md
# MAGIC # Gold: Análise de Subnotificações
# MAGIC 
# MAGIC Detecta possíveis subnotificações de óbito fetal cruzando laudos positivos e CIDs diagnósticos.
# MAGIC Identifica vínculos mãe-feto e valida contra auditoria oficial.

# COMMAND ----------

# MAGIC %run /Workspace/Libraries/Lake

# COMMAND ----------

from pyspark.sql import functions as F
from pyspark.sql import types as T
from pyspark.sql.utils import AnalysisException
from datetime import datetime
import pandas as pd

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# COMMAND ----------

# MAGIC %md
# MAGIC ## 1. Configuração

# COMMAND ----------

# Período analisado
PERIODO_INICIO = '2025-10-01'
PERIODO_FIM = '2025-11-01'

# Janela temporal (±dias) para buscar atendimentos relacionados
JANELA_DIAS = 7

# Tabelas
SILVER_TABLE = "innovation_dev.silver.auditoria_obitos_fetais_processado"
BRONZE_CID_TABLE = "innovation_dev.bronze.auditoria_obitos_cids"
AUDITORIA_TABLE = "RAWZN.RAW_HSP_TB_AUDITORIA_OBITO"

# Controle de camadas
FORCAR_REPROCESSAMENTO_CID = True  # True para forçar nova extração dos CIDs (CIDs corrigidos sem ponto)

# CIDs associados a óbito fetal (SEM PONTO - formato do Lake)
CID10_LIST = [
    # Núcleo (altamente específicos)
    'P95',      # Morte fetal de causa não especificada
    'P964',     # Morte neonatal precoce de causa não especificada
    'O364',     # Cuidado materno por morte intrauterina
    'O311',     # Morte de um feto ou mais em gestação múltipla
    'O312',     # Feto papiráceo
    'Z371',     # Nascimento de feto morto único
    'Z373',     # Gêmeos – um vivo e um morto
    'Z374',     # Gêmeos – ambos mortos
    'Z376',     # Múltiplos – alguns vivos e outros mortos
    'Z377',     # Múltiplos – todos mortos
    # Contexto forte (placenta/cordão)
    'O431',     # Descolamento prematuro da placenta
    'O691',     # Compressão do cordão umbilical
    'O692',     # Prolapso do cordão umbilical
    'O693',     # Circular de cordão com compressão
    'O698',     # Outras complicações do cordão
    'O699',     # Complicação não especificada do cordão
]

# Output
OUTPUT_PATH = "/Workspace/Innovation/t_eduardo.caminha/auditoria-obitos-fetais/outputs"
DATA_PROCESSAMENTO = datetime.now().strftime('%Y%m%d_%H%M%S')
EXCEL_FILENAME = f"subnotificacoes_{DATA_PROCESSAMENTO}.xlsx"
EXCEL_PATH = f"{OUTPUT_PATH}/{EXCEL_FILENAME}"

print("=" * 80)
print("CONFIGURAÇÃO - ANÁLISE DE SUBNOTIFICAÇÕES")
print("=" * 80)
print(f"Período: {PERIODO_INICIO} a {PERIODO_FIM}")
print(f"Janela temporal: ±{JANELA_DIAS} dias")
print(f"Total de CIDs monitorados: {len(CID10_LIST)}")
print(f"Tabela Silver: {SILVER_TABLE}")
print(f"Tabela Auditoria: {AUDITORIA_TABLE}")
print(f"Arquivo de saída: {EXCEL_FILENAME}")
print("=" * 80)

# COMMAND ----------

# MAGIC %md
# MAGIC ## 2. Conectar ao Datalake

# COMMAND ----------

connect_to_datalake(
    username="USR_PROD_INFORMATICA_SAUDE",
    password=dbutils.secrets.get(scope="INNOVATION_RAW", key="USR_PROD_INFORMATICA_SAUDE"),
    layer="RAWZN",
    level="LOW",
    dbx_secret_scope="INNOVATION_RAW"
)

print("✅ Conexão com datalake estabelecida")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 3. Ler Laudos Positivos (Silver)

# COMMAND ----------

try:
    df_laudos_silver = spark.table(SILVER_TABLE)
    
    # Aplicar filtro de período
    df_laudos_silver = df_laudos_silver.filter(
        (F.col('dt_procedimento_realizado') >= F.lit(PERIODO_INICIO)) &
        (F.col('dt_procedimento_realizado') < F.lit(PERIODO_FIM))
    )
    
    total_laudos = df_laudos_silver.count()
    print(f"✅ Laudos positivos carregados: {total_laudos:,}")
    
    if total_laudos == 0:
        raise RuntimeError("⚠️ Nenhum laudo positivo encontrado para o período. Execute o processamento Silver primeiro.")
    
except AnalysisException:
    raise RuntimeError(f"⚠️ Tabela Silver {SILVER_TABLE} não encontrada. Execute o processamento Silver primeiro.")

df_laudos_silver.createOrReplaceTempView("vw_laudos_positivos")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 4. FONTE 1: Vincular Mães e Fetos dos Laudos

# COMMAND ----------

print("🔍 Processando vínculos dos LAUDOS...")

df_laudos_pd = df_laudos_silver.select(
    'cd_paciente', 'nm_paciente', 'dt_procedimento_realizado', 'fonte', 'cd_atendimento'
).distinct().toPandas()

vinculos_laudos_list = []

for idx, laudo in df_laudos_pd.iterrows():
    if idx % 10 == 0:
        print(f"   Processando laudo {idx+1}/{len(df_laudos_pd)}...")
    
    cd_paciente_mae = laudo['cd_paciente']
    nm_mae = laudo['nm_paciente']
    dt_ref = laudo['dt_procedimento_realizado']
    
    # Calcular janela ±7 dias
    dt_inicio = (pd.to_datetime(dt_ref) - pd.Timedelta(days=JANELA_DIAS)).strftime('%Y-%m-%d')
    dt_fim = (pd.to_datetime(dt_ref) + pd.Timedelta(days=JANELA_DIAS)).strftime('%Y-%m-%d')
    
    # 1. Buscar TODOS os atendimentos da mãe na janela (Lake)
    query_atend_mae = f"""
    SELECT CD_ATENDIMENTO, CD_PACIENTE, DT_ATENDIMENTO
    FROM RAWZN.RAW_HSP_TM_ATENDIMENTO
    WHERE CD_PACIENTE = {cd_paciente_mae}
      AND DT_ATENDIMENTO >= DATE '{dt_inicio}'
      AND DT_ATENDIMENTO <= DATE '{dt_fim}'
    UNION ALL
    SELECT CD_ATENDIMENTO, CD_PACIENTE, DT_ATENDIMENTO
    FROM RAWZN.RAW_PSC_TM_ATENDIMENTO
    WHERE CD_PACIENTE = {cd_paciente_mae}
      AND DT_ATENDIMENTO >= DATE '{dt_inicio}'
      AND DT_ATENDIMENTO <= DATE '{dt_fim}'
    """
    
    atend_mae_pd = run_sql(query_atend_mae)
    
    if len(atend_mae_pd) > 0:
        cd_atend_maes = ", ".join(str(int(x)) for x in atend_mae_pd['CD_ATENDIMENTO'].unique())
        
        # 2. Buscar fetos vinculados a esses atendimentos (CD_ATENDIMENTO_MAE)
        query_fetos = f"""
        SELECT CD_ATENDIMENTO, CD_PACIENTE, CD_ATENDIMENTO_MAE, DT_ATENDIMENTO
        FROM RAWZN.RAW_HSP_TM_ATENDIMENTO
        WHERE CD_ATENDIMENTO_MAE IN ({cd_atend_maes})
        UNION ALL
        SELECT CD_ATENDIMENTO, CD_PACIENTE, CD_ATENDIMENTO_MAE, DT_ATENDIMENTO
        FROM RAWZN.RAW_PSC_TM_ATENDIMENTO
        WHERE CD_ATENDIMENTO_MAE IN ({cd_atend_maes})
        """
        
        fetos_pd = run_sql(query_fetos)
        
        if len(fetos_pd) > 0:
            # Tem fetos vinculados
            for _, feto in fetos_pd.iterrows():
                vinculos_laudos_list.append({
                    'cd_paciente_principal': cd_paciente_mae,
                    'nm_paciente_principal': nm_mae,
                    'cd_paciente_feto': int(feto['CD_PACIENTE']),
                    'tem_feto_vinculado': 'SIM',
                    'origem': 'LAUDO'
                })
        else:
            # Não tem fetos vinculados (possível subnotificação)
            vinculos_laudos_list.append({
                'cd_paciente_principal': cd_paciente_mae,
                'nm_paciente_principal': nm_mae,
                'cd_paciente_feto': None,
                'tem_feto_vinculado': 'NAO',
                'origem': 'LAUDO'
            })

if len(vinculos_laudos_list) > 0:
    df_laudos_vinculos = spark.createDataFrame(pd.DataFrame(vinculos_laudos_list))
else:
    schema = T.StructType([
        T.StructField('cd_paciente_principal', T.LongType(), True),
        T.StructField('nm_paciente_principal', T.StringType(), True),
        T.StructField('cd_paciente_feto', T.LongType(), True),
        T.StructField('tem_feto_vinculado', T.StringType(), True),
        T.StructField('origem', T.StringType(), True)
    ])
    df_laudos_vinculos = spark.createDataFrame([], schema)

total_vinculos_laudos = df_laudos_vinculos.count()
com_feto = df_laudos_vinculos.filter(F.col('cd_paciente_feto').isNotNull()).count()
sem_feto = total_vinculos_laudos - com_feto

print("=" * 80)
print("LAUDOS - VÍNCULOS IDENTIFICADOS")
print("=" * 80)
print(f"Total de casos: {total_vinculos_laudos:,}")
print(f"Com feto vinculado: {com_feto:,}")
print(f"Sem feto vinculado: {sem_feto:,}")
print("=" * 80)

df_laudos_vinculos.createOrReplaceTempView("vw_laudos_vinculos")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 5. FONTE 2: Buscar CIDs de Óbito Fetal (com Bronze)

# COMMAND ----------

def carregar_cids_bronze():
    """Tenta carregar CIDs da Bronze. Se não existir ou forçar, extrai do Lake."""
    if FORCAR_REPROCESSAMENTO_CID:
        print("⚠️ Forçando reprocessamento dos CIDs")
        return None
    
    try:
        cids_bronze = spark.table(BRONZE_CID_TABLE)
        total_bronze = cids_bronze.count()
        print(f"✅ CIDs carregados da Bronze: {total_bronze:,}")
        return cids_bronze
    except AnalysisException:
        print(f"ℹ️ Tabela Bronze CID {BRONZE_CID_TABLE} não encontrada. Extraindo do Lake...")
        return None

df_cids = carregar_cids_bronze()

if df_cids is None:
    # Extrair do Lake usando run_sql
    cid_sql_list = ", ".join(f"'{cid}'" for cid in CID10_LIST)
    
    query_cids = f"""
    SELECT
        'HSP' AS FONTE,
        D.CD_ATENDIMENTO,
        A.CD_PACIENTE,
        D.CD_CID10,
        D.DT_DIAGNOSTICO AS DT_REFERENCIA,
        NVL(D.FL_VALIDADO, 'S') AS FL_VALIDADO
    FROM RAWZN.RAW_HSP_TB_DIAGNOSTICO_ATENDIMENTO D
    INNER JOIN RAWZN.RAW_HSP_TM_ATENDIMENTO A
        ON D.CD_ATENDIMENTO = A.CD_ATENDIMENTO
    WHERE D.CD_CID10 IN ({cid_sql_list})
      AND D.DT_DIAGNOSTICO >= DATE '{PERIODO_INICIO}'
      AND D.DT_DIAGNOSTICO < DATE '{PERIODO_FIM}'
    
    UNION ALL
    
    SELECT
        'PSC' AS FONTE,
        D.CD_ATENDIMENTO,
        A.CD_PACIENTE,
        D.CD_CID10,
        D.DT_DIAGNOSTICO AS DT_REFERENCIA,
        NVL(D.FL_VALIDADO, 'S') AS FL_VALIDADO
    FROM RAWZN.RAW_PSC_TB_DIAGNOSTICO_ATENDIMENTO D
    INNER JOIN RAWZN.RAW_PSC_TM_ATENDIMENTO A
        ON D.CD_ATENDIMENTO = A.CD_ATENDIMENTO
    WHERE D.CD_CID10 IN ({cid_sql_list})
      AND D.DT_DIAGNOSTICO >= DATE '{PERIODO_INICIO}'
      AND D.DT_DIAGNOSTICO < DATE '{PERIODO_FIM}'
    """
    
    cids_pd = run_sql(query_cids)
    
    # Filtrar validados
    cids_pd = cids_pd[cids_pd['FL_VALIDADO'] == 'S'].copy()
    cids_pd = cids_pd.drop(columns=['FL_VALIDADO'])
    
    if len(cids_pd) > 0:
        # Converter tipos
        cids_pd['DT_REFERENCIA'] = pd.to_datetime(cids_pd['DT_REFERENCIA'])
        
        df_cids = spark.createDataFrame(cids_pd)
        df_cids = df_cids.withColumn("DT_INGESTAO", F.current_timestamp())
        
        # Salvar em Bronze
        print(f"💾 Gravando {len(cids_pd):,} CIDs na Bronze: {BRONZE_CID_TABLE}")
        df_cids.write.format("delta").mode("overwrite").option("overwriteSchema", "true").saveAsTable(BRONZE_CID_TABLE)
        spark.catalog.refreshTable(BRONZE_CID_TABLE)
    else:
        print("⚠️ Nenhum CID encontrado para o período")
        df_cids = spark.createDataFrame([], T.StructType([
            T.StructField('FONTE', T.StringType(), True),
            T.StructField('CD_ATENDIMENTO', T.LongType(), True),
            T.StructField('CD_PACIENTE', T.LongType(), True),
            T.StructField('CD_CID10', T.StringType(), True),
            T.StructField('DT_REFERENCIA', T.TimestampType(), True)
        ]))

# Aplicar filtro de período (caso tenha carregado Bronze com período maior)
df_cids = df_cids.filter(
    (F.col('DT_REFERENCIA') >= F.lit(PERIODO_INICIO)) &
    (F.col('DT_REFERENCIA') < F.lit(PERIODO_FIM))
)

total_cids = df_cids.count()
print(f"✅ Diagnósticos com CIDs de óbito fetal: {total_cids:,}")

if total_cids > 0:
    print(f"\n📋 CIDs mais frequentes:")
    df_cids.groupBy('CD_CID10').count().orderBy(F.desc('count')).show(10, truncate=False)

df_cids.createOrReplaceTempView("vw_cids_obito")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 6. FONTE 2: Vincular Pacientes com CID e seus Fetos

# COMMAND ----------

print("🔍 Processando vínculos dos CIDs...")

df_cids_pd = df_cids.select('CD_PACIENTE', 'CD_ATENDIMENTO', 'DT_REFERENCIA', 'CD_CID10').distinct().toPandas()

vinculos_cids_list = []

for idx, cid_row in df_cids_pd.iterrows():
    if idx % 10 == 0:
        print(f"   Processando CID {idx+1}/{len(df_cids_pd)}...")
    
    cd_paciente = int(cid_row['CD_PACIENTE'])
    cd_atendimento_cid = int(cid_row['CD_ATENDIMENTO'])
    dt_ref = cid_row['DT_REFERENCIA']
    cd_cid10 = cid_row['CD_CID10']
    
    # Buscar nome do paciente
    query_nome_paciente = f"""
    SELECT NM_PACIENTE
    FROM RAWZN.RAW_HSP_TB_PACIENTE
    WHERE CD_PACIENTE = {cd_paciente}
    UNION ALL
    SELECT NM_PACIENTE
    FROM RAWZN.RAW_PSC_TB_PACIENTE
    WHERE CD_PACIENTE = {cd_paciente}
    """
    
    try:
        nome_result = run_sql(query_nome_paciente)
        nm_paciente = nome_result['NM_PACIENTE'].iloc[0] if len(nome_result) > 0 else None
    except:
        nm_paciente = None
    
    # 1. Buscar o atendimento específico do CID para ver se tem CD_ATENDIMENTO_MAE
    query_atend_cid = f"""
    SELECT CD_ATENDIMENTO, CD_PACIENTE, CD_ATENDIMENTO_MAE
    FROM RAWZN.RAW_HSP_TM_ATENDIMENTO
    WHERE CD_ATENDIMENTO = {cd_atendimento_cid}
    UNION ALL
    SELECT CD_ATENDIMENTO, CD_PACIENTE, CD_ATENDIMENTO_MAE
    FROM RAWZN.RAW_PSC_TM_ATENDIMENTO
    WHERE CD_ATENDIMENTO = {cd_atendimento_cid}
    """
    
    atend_cid_pd = run_sql(query_atend_cid)
    
    if len(atend_cid_pd) > 0:
        cd_atendimento_mae_col = atend_cid_pd.iloc[0]['CD_ATENDIMENTO_MAE']
        
        # CASO 1: CD_ATENDIMENTO_MAE = NULL → Paciente principal (buscar fetos dele)
        if pd.isna(cd_atendimento_mae_col):
            # Calcular janela ±7 dias
            dt_inicio = (pd.to_datetime(dt_ref) - pd.Timedelta(days=JANELA_DIAS)).strftime('%Y-%m-%d')
            dt_fim = (pd.to_datetime(dt_ref) + pd.Timedelta(days=JANELA_DIAS)).strftime('%Y-%m-%d')
            
            # Buscar todos atendimentos deste paciente na janela
            query_atend_principal = f"""
            SELECT CD_ATENDIMENTO
            FROM RAWZN.RAW_HSP_TM_ATENDIMENTO
            WHERE CD_PACIENTE = {cd_paciente}
              AND DT_ATENDIMENTO >= DATE '{dt_inicio}'
              AND DT_ATENDIMENTO <= DATE '{dt_fim}'
            UNION ALL
            SELECT CD_ATENDIMENTO
            FROM RAWZN.RAW_PSC_TM_ATENDIMENTO
            WHERE CD_PACIENTE = {cd_paciente}
              AND DT_ATENDIMENTO >= DATE '{dt_inicio}'
              AND DT_ATENDIMENTO <= DATE '{dt_fim}'
            """
            
            atend_principal_pd = run_sql(query_atend_principal)
            
            if len(atend_principal_pd) > 0:
                cd_atends = ", ".join(str(int(x)) for x in atend_principal_pd['CD_ATENDIMENTO'].unique())
                
                # Buscar fetos vinculados
                query_fetos = f"""
                SELECT CD_ATENDIMENTO, CD_PACIENTE, CD_ATENDIMENTO_MAE
                FROM RAWZN.RAW_HSP_TM_ATENDIMENTO
                WHERE CD_ATENDIMENTO_MAE IN ({cd_atends})
                UNION ALL
                SELECT CD_ATENDIMENTO, CD_PACIENTE, CD_ATENDIMENTO_MAE
                FROM RAWZN.RAW_PSC_TM_ATENDIMENTO
                WHERE CD_ATENDIMENTO_MAE IN ({cd_atends})
                """
                
                fetos_pd = run_sql(query_fetos)
                
                if len(fetos_pd) > 0:
                    # Tem fetos vinculados
                    for _, feto in fetos_pd.iterrows():
                        vinculos_cids_list.append({
                            'cd_paciente_principal': cd_paciente,
                            'nm_paciente_principal': nm_paciente,
                            'cd_paciente_feto': int(feto['CD_PACIENTE']),
                            'tem_feto_vinculado': 'SIM',
                            'origem': 'CID'
                        })
                else:
                    # Não tem fetos vinculados
                    vinculos_cids_list.append({
                        'cd_paciente_principal': cd_paciente,
                        'nm_paciente_principal': nm_paciente,
                        'cd_paciente_feto': None,
                        'tem_feto_vinculado': 'NAO',
                        'origem': 'CID'
                    })
        
        # CASO 2: CD_ATENDIMENTO_MAE != NULL → Este é um FETO, buscar a mãe
        else:
            cd_atendimento_mae = int(cd_atendimento_mae_col)
            
            # Buscar CD_PACIENTE da mãe através do CD_ATENDIMENTO_MAE
            query_mae = f"""
            SELECT CD_ATENDIMENTO, CD_PACIENTE
            FROM RAWZN.RAW_HSP_TM_ATENDIMENTO
            WHERE CD_ATENDIMENTO = {cd_atendimento_mae}
            UNION ALL
            SELECT CD_ATENDIMENTO, CD_PACIENTE
            FROM RAWZN.RAW_PSC_TM_ATENDIMENTO
            WHERE CD_ATENDIMENTO = {cd_atendimento_mae}
            """
            
            mae_pd = run_sql(query_mae)
            
            if len(mae_pd) > 0:
                cd_paciente_mae = int(mae_pd.iloc[0]['CD_PACIENTE'])
                
                # Buscar nome da mãe em TB_PACIENTE
                query_nome_mae = f"""
                SELECT NM_PACIENTE
                FROM RAWZN.RAW_HSP_TB_PACIENTE
                WHERE CD_PACIENTE = {cd_paciente_mae}
                UNION ALL
                SELECT NM_PACIENTE
                FROM RAWZN.RAW_PSC_TB_PACIENTE
                WHERE CD_PACIENTE = {cd_paciente_mae}
                """
                
                try:
                    nome_mae_result = run_sql(query_nome_mae)
                    nm_paciente_mae = nome_mae_result['NM_PACIENTE'].iloc[0] if len(nome_mae_result) > 0 else None
                except:
                    nm_paciente_mae = None
                
                vinculos_cids_list.append({
                    'cd_paciente_principal': cd_paciente_mae,
                    'nm_paciente_principal': nm_paciente_mae,
                    'cd_paciente_feto': cd_paciente,
                    'tem_feto_vinculado': 'SIM',
                    'origem': 'CID'
                })

if len(vinculos_cids_list) > 0:
    df_cids_vinculos = spark.createDataFrame(pd.DataFrame(vinculos_cids_list))
else:
    schema = T.StructType([
        T.StructField('cd_paciente_principal', T.LongType(), True),
        T.StructField('nm_paciente_principal', T.StringType(), True),
        T.StructField('cd_paciente_feto', T.LongType(), True),
        T.StructField('tem_feto_vinculado', T.StringType(), True),
        T.StructField('origem', T.StringType(), True)
    ])
    df_cids_vinculos = spark.createDataFrame([], schema)

total_vinculos_cids = df_cids_vinculos.count()
com_feto_cid = df_cids_vinculos.filter(F.col('cd_paciente_feto').isNotNull()).count()
sem_feto_cid = total_vinculos_cids - com_feto_cid

print("=" * 80)
print("CIDs - VÍNCULOS IDENTIFICADOS")
print("=" * 80)
print(f"Total de casos: {total_vinculos_cids:,}")
print(f"Com feto vinculado: {com_feto_cid:,}")
print(f"Sem feto vinculado: {sem_feto_cid:,}")
print("=" * 80)

df_cids_vinculos.createOrReplaceTempView("vw_cids_vinculos")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 7. União e Deduplicação

# COMMAND ----------

# União das duas fontes (já têm a mesma estrutura)
df_uniao = df_laudos_vinculos.union(df_cids_vinculos)

# Agregar por (cd_paciente_principal, cd_paciente_feto) para identificar origem combinada
df_consolidado = df_uniao.groupBy(
    'cd_paciente_principal', 'cd_paciente_feto'
).agg(
    F.max('nm_paciente_principal').alias('nm_paciente_principal'),
    F.max('tem_feto_vinculado').alias('tem_feto_vinculado'),
    F.collect_set('origem').alias('origens')
)

# Criar flag de origem combinada
df_consolidado = df_consolidado.withColumn(
    'origem_deteccao',
    F.when(F.array_contains('origens', 'LAUDO') & F.array_contains('origens', 'CID'), F.lit('AMBOS'))
     .when(F.array_contains('origens', 'LAUDO'), F.lit('LAUDO'))
     .otherwise(F.lit('CID'))
).drop('origens')

total_pares = df_consolidado.count()
com_feto_total = df_consolidado.filter(F.col('cd_paciente_feto').isNotNull()).count()
sem_feto_total = total_pares - com_feto_total

print("=" * 80)
print("CONSOLIDAÇÃO FINAL")
print("=" * 80)
print(f"Total de casos únicos: {total_pares:,}")
print(f"Com feto vinculado: {com_feto_total:,}")
print(f"Sem feto vinculado: {sem_feto_total:,}")
print("\n📊 Por origem de detecção:")
df_consolidado.groupBy('origem_deteccao').count().orderBy('origem_deteccao').show()
print("=" * 80)

df_consolidado.createOrReplaceTempView("vw_pares_consolidados")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 8. Buscar Nomes dos Pacientes e Fetos

# COMMAND ----------

df_consolidado_pd = df_consolidado.toPandas()

# Identificar pacientes principais sem nome
pacientes_sem_nome = df_consolidado_pd[df_consolidado_pd['nm_paciente_principal'].isna()]['cd_paciente_principal'].unique()

if len(pacientes_sem_nome) > 0:
    pacientes_csv = ", ".join(str(int(x)) for x in pacientes_sem_nome if pd.notna(x))
    
    if pacientes_csv:
        print(f"🔍 Buscando nomes de {len(pacientes_sem_nome)} pacientes principais...")
        
        # Tentar primeiro em TB_PACIENTE
        query_nomes = f"""
        SELECT CD_PACIENTE, NM_PACIENTE
        FROM RAWZN.RAW_HSP_TB_PACIENTE
        WHERE CD_PACIENTE IN ({pacientes_csv})
        UNION ALL
        SELECT CD_PACIENTE, NM_PACIENTE
        FROM RAWZN.RAW_PSC_TB_PACIENTE
        WHERE CD_PACIENTE IN ({pacientes_csv})
        """
        
        try:
            nomes_pd = run_sql(query_nomes)
            nomes_map = dict(zip(nomes_pd['CD_PACIENTE'], nomes_pd['NM_PACIENTE']))
            print(f"   ✅ {len(nomes_map)} nomes encontrados em TB_PACIENTE")
        except Exception as e:
            print(f"   ⚠️  Erro ao buscar em TB_PACIENTE: {e}")
            nomes_map = {}
        
        
        # Mapear nomes dos pacientes principais
        df_consolidado_pd['nm_paciente_principal'] = df_consolidado_pd.apply(
            lambda row: nomes_map.get(row['cd_paciente_principal'], row['nm_paciente_principal']) 
                        if pd.isna(row['nm_paciente_principal']) else row['nm_paciente_principal'],
            axis=1
        )
        
        # Verificar quantos ainda estão sem nome
        ainda_vazios = df_consolidado_pd['nm_paciente_principal'].isna().sum()
        if ainda_vazios > 0:
            print(f"   ⚠️  {ainda_vazios} pacientes principais ainda sem nome após busca")

# Buscar nomes dos fetos
fetos_com_cd = df_consolidado_pd[df_consolidado_pd['cd_paciente_feto'].notna()]['cd_paciente_feto'].unique()

if len(fetos_com_cd) > 0:
    fetos_csv = ", ".join(str(int(x)) for x in fetos_com_cd)
    
    print(f"🔍 Buscando nomes de {len(fetos_com_cd)} fetos...")
    
    # Tentar primeiro em TB_PACIENTE
    query_nomes_fetos = f"""
    SELECT CD_PACIENTE, NM_PACIENTE
    FROM RAWZN.RAW_HSP_TB_PACIENTE
    WHERE CD_PACIENTE IN ({fetos_csv})
    UNION ALL
    SELECT CD_PACIENTE, NM_PACIENTE
    FROM RAWZN.RAW_PSC_TB_PACIENTE
    WHERE CD_PACIENTE IN ({fetos_csv})
    """
    
    try:
        nomes_fetos_pd = run_sql(query_nomes_fetos)
        nomes_fetos_map = dict(zip(nomes_fetos_pd['CD_PACIENTE'], nomes_fetos_pd['NM_PACIENTE']))
        print(f"   ✅ {len(nomes_fetos_map)} nomes de fetos encontrados em TB_PACIENTE")
    except Exception as e:
        print(f"   ⚠️  Erro ao buscar em TB_PACIENTE: {e}")
        nomes_fetos_map = {}
    
    
    # Mapear nomes dos fetos
    df_consolidado_pd['nm_paciente_feto'] = df_consolidado_pd['cd_paciente_feto'].apply(
        lambda x: nomes_fetos_map.get(int(x)) if pd.notna(x) else None
    )
    
    # Verificar quantos fetos ainda estão sem nome
    fetos_vazios = df_consolidado_pd[df_consolidado_pd['cd_paciente_feto'].notna()]['nm_paciente_feto'].isna().sum()
    if fetos_vazios > 0:
        print(f"   ⚠️  {fetos_vazios} fetos ainda sem nome após busca")
else:
    df_consolidado_pd['nm_paciente_feto'] = None

df_consolidado_nomes = spark.createDataFrame(df_consolidado_pd)

print("✅ Nomes dos pacientes complementados")

df_consolidado_nomes.createOrReplaceTempView("vw_pares_com_nomes")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 9. Checar na Auditoria Oficial (RAW_HSP_TB_AUDITORIA_OBITO)

# COMMAND ----------

# Coletar todos CD_ATENDIMENTO dos casos detectados (Laudos + CIDs)
# Vou buscar apenas esses específicos na auditoria (muito mais eficiente)

# CD_ATENDIMENTO dos laudos
atendimentos_laudos = set(
    df_laudos_pd[df_laudos_pd['cd_atendimento'].notna()]['cd_atendimento'].astype(str).values
)

# CD_ATENDIMENTO dos CIDs
atendimentos_cids = set(
    cids_pd[cids_pd['CD_ATENDIMENTO'].notna()]['CD_ATENDIMENTO'].astype(str).values
)

# União de todos os CD_ATENDIMENTO detectados
todos_atendimentos = atendimentos_laudos.union(atendimentos_cids)

print(f"📋 Total de CD_ATENDIMENTO a verificar na auditoria: {len(todos_atendimentos):,}")
print(f"   - Laudos: {len(atendimentos_laudos):,}")
print(f"   - CIDs: {len(atendimentos_cids):,}")

# Criar lista para IN clause (limitada para evitar query muito grande)
MAX_BATCH = 1000
batches = [list(todos_atendimentos)[i:i+MAX_BATCH] for i in range(0, len(todos_atendimentos), MAX_BATCH)]

atendimentos_na_auditoria = set()

for batch_idx, batch in enumerate(batches, 1):
    cd_atendimentos_str = ','.join(batch)
    
    query_batch = f"""
    SELECT DISTINCT CD_ATENDIMENTO
    FROM RAWZN.RAW_HSP_TB_AUDITORIA_OBITO
    WHERE CD_ATENDIMENTO IN ({cd_atendimentos_str})
    """
    
    try:
        batch_result = run_sql(query_batch)
        atendimentos_na_auditoria.update(batch_result['CD_ATENDIMENTO'].astype(str).values)
        print(f"   Batch {batch_idx}/{len(batches)}: {len(batch_result)} atendimentos encontrados")
    except Exception as e:
        print(f"   ⚠️  Erro no batch {batch_idx}: {str(e)}")

print(f"✅ Total de CD_ATENDIMENTO encontrados na auditoria: {len(atendimentos_na_auditoria):,}")

# Marcar atendimentos dos casos detectados que estão na auditoria
# Laudos
df_laudos_pd['atendimento_na_auditoria'] = df_laudos_pd['cd_atendimento'].astype(str).apply(
    lambda x: x in atendimentos_na_auditoria
)

# CIDs
cids_pd['atendimento_na_auditoria'] = cids_pd['CD_ATENDIMENTO'].astype(str).apply(
    lambda x: x in atendimentos_na_auditoria
)

# Agora verifico: cada caso no consolidado está na auditoria?
# Um caso está na auditoria se seu CD_ATENDIMENTO original (laudo ou CID) está lá
df_consolidado_nomes_pd = df_consolidado_nomes.toPandas()

# Para cada linha do consolidado, verifico se veio de um atendimento auditado
def verificar_auditoria(row):
    cd_principal = str(row['cd_paciente_principal'])
    origem = row['origem_deteccao']
    
    # Verificar nos laudos
    if 'LAUDO' in origem:
        laudos_match = df_laudos_pd[df_laudos_pd['cd_paciente'].astype(str) == cd_principal]
        if not laudos_match.empty and laudos_match['atendimento_na_auditoria'].any():
            return 'SIM'
    
    # Verificar nos CIDs
    if 'CID' in origem:
        cids_match = cids_pd[cids_pd['CD_PACIENTE'].astype(str) == cd_principal]
        if not cids_match.empty and cids_match['atendimento_na_auditoria'].any():
            return 'SIM'
    
    return 'NAO'

df_consolidado_nomes_pd['na_auditoria'] = df_consolidado_nomes_pd.apply(verificar_auditoria, axis=1)

df_com_auditoria = spark.createDataFrame(df_consolidado_nomes_pd)

total_na_auditoria = df_com_auditoria.filter(F.col('na_auditoria') == 'SIM').count()
total_nao_auditados = df_com_auditoria.filter(F.col('na_auditoria') == 'NAO').count()

print("=" * 80)
print("CHECAGEM NA AUDITORIA OFICIAL")
print("=" * 80)
print(f"Total de casos: {df_com_auditoria.count():,}")
print(f"Presentes na auditoria: {total_na_auditoria:,}")
print(f"NÃO presentes (subnotificações): {total_nao_auditados:,}")
print("=" * 80)

# COMMAND ----------

# MAGIC %md
# MAGIC ## 10. Enriquecer com Laudos e CIDs

# COMMAND ----------

print("🔍 Enriquecendo dados com laudos e CIDs...")

df_enriquecido_pd = df_com_auditoria.toPandas()

# Buscar laudos (apenas o mais recente para cada mãe)
laudos_silver_pd = df_laudos_silver.select(
    'cd_paciente', 'dt_procedimento_realizado', 'texto_original', 'termo_detectado'
).toPandas()

# Para cada paciente principal, buscar o laudo mais recente
def buscar_ultimo_laudo(cd_paciente):
    if pd.isna(cd_paciente):
        return None
    
    laudos_paciente = laudos_silver_pd[laudos_silver_pd['cd_paciente'] == cd_paciente]
    if len(laudos_paciente) == 0:
        return None
    
    # Ordenar por data e pegar o mais recente
    laudos_paciente = laudos_paciente.sort_values('dt_procedimento_realizado', ascending=False)
    laudo_recente = laudos_paciente.iloc[0]['texto_original']
    
    # Retornar laudo completo
    return laudo_recente if laudo_recente else None

# Buscar CIDs
cids_bronze_pd = df_cids.select('CD_PACIENTE', 'CD_CID10').toPandas()

# Garantir tipos consistentes para comparação
cids_bronze_pd['CD_PACIENTE'] = cids_bronze_pd['CD_PACIENTE'].astype('Int64')

def buscar_cids(cd_paciente):
    if pd.isna(cd_paciente):
        return None
    
    # Converter para Int64 para garantir match
    try:
        cd_paciente_int = int(cd_paciente)
    except (ValueError, TypeError):
        return None
    
    cids_paciente = cids_bronze_pd[cids_bronze_pd['CD_PACIENTE'] == cd_paciente_int]
    if len(cids_paciente) == 0:
        return None
    
    # Retornar lista única de CIDs
    cids_lista = cids_paciente['CD_CID10'].unique().tolist()
    return ", ".join(cids_lista)

print(f"   📊 CIDs disponíveis na Bronze: {len(cids_bronze_pd)} registros")
print(f"   📊 Pacientes únicos na Bronze CID: {cids_bronze_pd['CD_PACIENTE'].nunique()}")

print("   Buscando laudos das mães...")
df_enriquecido_pd['laudo_mae'] = df_enriquecido_pd['cd_paciente_principal'].apply(buscar_ultimo_laudo)

print("   Buscando CIDs das mães...")
df_enriquecido_pd['cids_mae'] = df_enriquecido_pd['cd_paciente_principal'].apply(buscar_cids)

print("   Buscando CIDs dos fetos...")
df_enriquecido_pd['cids_feto'] = df_enriquecido_pd['cd_paciente_feto'].apply(buscar_cids)

# Debug: quantos encontramos
cids_mae_encontrados = df_enriquecido_pd['cids_mae'].notna().sum()
cids_feto_encontrados = df_enriquecido_pd['cids_feto'].notna().sum()
print(f"   ✅ CIDs encontrados: {cids_mae_encontrados} mães, {cids_feto_encontrados} fetos")

# Converter de volta para Spark
df_com_auditoria = spark.createDataFrame(df_enriquecido_pd)

print("✅ Dados enriquecidos com laudos e CIDs")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 11. Filtrar Apenas Subnotificações

# COMMAND ----------

df_subnotificacoes = df_com_auditoria.filter(F.col('na_auditoria') == 'NAO')

print(f"🚨 SUBNOTIFICAÇÕES DETECTADAS: {df_subnotificacoes.count():,}")

if df_subnotificacoes.count() > 0:
    print("\n📊 Por origem de detecção:")
    df_subnotificacoes.groupBy('origem_deteccao').count().orderBy('origem_deteccao').show()
    
    print("\n📊 Com/sem feto identificado:")
    df_subnotificacoes.withColumn(
        'feto_identificado',
        F.when(F.col('cd_paciente_feto').isNotNull(), 'COM_FETO').otherwise('SEM_FETO')
    ).groupBy('feto_identificado').count().show()

# COMMAND ----------

# MAGIC %md
# MAGIC ## 12. Exportar para Excel

# COMMAND ----------

# Converter para pandas
subnotificacoes_pd = df_subnotificacoes.toPandas()
todos_casos_pd = df_com_auditoria.toPandas()

# Reordenar colunas para melhor visualização
colunas_ordenadas = [
    'cd_paciente_principal',
    'nm_paciente_principal',
    'cd_paciente_feto',
    'nm_paciente_feto',
    'tem_feto_vinculado',
    'origem_deteccao',
    'laudo_mae',
    'cids_mae',
    'cids_feto',
    'na_auditoria'
]

# Aplicar reordenação (se todas as colunas existirem)
colunas_existentes = [col for col in colunas_ordenadas if col in subnotificacoes_pd.columns]
subnotificacoes_pd = subnotificacoes_pd[colunas_existentes]
todos_casos_pd = todos_casos_pd[colunas_existentes]

# Estatísticas
stats_origem = todos_casos_pd.groupby(['origem_deteccao', 'na_auditoria']).size().reset_index(name='total')
stats_feto = todos_casos_pd.groupby([
    'na_auditoria',
    todos_casos_pd['cd_paciente_feto'].notna().map({True: 'COM_FETO', False: 'SEM_FETO'})
]).size().reset_index(name='total')

stats_overview = pd.DataFrame({
    'metrica': [
        'Total de casos detectados',
        'Casos na auditoria',
        'SUBNOTIFICAÇÕES DETECTADAS',
        'Subnotificações com feto identificado',
        'Subnotificações sem feto identificado',
        'Detectados por LAUDO apenas',
        'Detectados por CID apenas',
        'Detectados por AMBOS',
        'Período analisado',
        'Janela temporal',
        'Data de processamento'
    ],
    'valor': [
        len(todos_casos_pd),
        len(todos_casos_pd[todos_casos_pd['na_auditoria'] == 'SIM']),
        len(subnotificacoes_pd),
        len(subnotificacoes_pd[subnotificacoes_pd['cd_paciente_feto'].notna()]),
        len(subnotificacoes_pd[subnotificacoes_pd['cd_paciente_feto'].isna()]),
        len(todos_casos_pd[todos_casos_pd['origem_deteccao'] == 'LAUDO']),
        len(todos_casos_pd[todos_casos_pd['origem_deteccao'] == 'CID']),
        len(todos_casos_pd[todos_casos_pd['origem_deteccao'] == 'AMBOS']),
        f"{PERIODO_INICIO} a {PERIODO_FIM}",
        f"±{JANELA_DIAS} dias",
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ]
})

# Criar Excel
wb = Workbook()
wb.remove(wb.active)

def escrever_aba(nome, dataframe):
    """Escreve um DataFrame em uma aba do Excel com formatação"""
    ws = wb.create_sheet(nome)
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        ws.append(row)
    
    # Formatar cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # Ajustar largura das colunas
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(adjusted_width, 60)

# Criar abas
escrever_aba("1_Resumo", stats_overview)
escrever_aba("2_Subnotificacoes", subnotificacoes_pd)
escrever_aba("3_Todos_Casos", todos_casos_pd)
escrever_aba("4_Stats_Origem", stats_origem)
escrever_aba("5_Stats_Feto", stats_feto)

# Salvar
import os
os.makedirs(OUTPUT_PATH, exist_ok=True)
wb.save(EXCEL_PATH)

print("=" * 80)
print("✅ ARQUIVO EXCEL GERADO COM SUCESSO!")
print("=" * 80)
print(f"📁 Local: {EXCEL_PATH}")
print(f"📊 Abas criadas:")
print(f"   1. Resumo Geral")
print(f"   2. Subnotificações ({len(subnotificacoes_pd):,} casos)")
print(f"   3. Todos os Casos ({len(todos_casos_pd):,} registros)")
print(f"   4. Estatísticas por Origem")
print(f"   5. Estatísticas Feto Identificado")
print(f"\n📋 Colunas incluídas:")
print(f"   • cd_paciente_principal / nm_paciente_principal")
print(f"   • cd_paciente_feto / nm_paciente_feto")
print(f"   • laudo_mae (último laudo do US)")
print(f"   • cids_mae (CIDs da mãe)")
print(f"   • cids_feto (CIDs do feto, quando houver)")
print(f"   • tem_feto_vinculado, origem_deteccao, na_auditoria")
print("=" * 80)

# COMMAND ----------

# MAGIC %md
# MAGIC ## 13. Resumo Final

# COMMAND ----------

print("=" * 80)
print("RESUMO FINAL - ANÁLISE DE SUBNOTIFICAÇÕES")
print("=" * 80)
print(f"🔍 Período analisado: {PERIODO_INICIO} a {PERIODO_FIM}")
print(f"📋 Laudos positivos: {total_laudos:,}")
print(f"🏥 Diagnósticos CID: {total_cids:,}")
print(f"👥 Pares (mãe, feto) únicos: {total_pares:,}")
print(f"✅ Casos na auditoria: {total_na_auditoria:,}")
print(f"🚨 SUBNOTIFICAÇÕES DETECTADAS: {total_nao_auditados:,}")
print(f"📁 Arquivo: {EXCEL_FILENAME}")
print("=" * 80)

